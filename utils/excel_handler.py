# utils/excel_handler.py
import openpyxl
import allure
import pandas as pd
from datetime import datetime
import os
from .config_handler import Configuration


class ExcelHandler:
    @staticmethod
    def load_workbook(file_path):
        """Load Excel workbook"""
        try:
            return openpyxl.load_workbook(file_path)
        except Exception as e:
            error_msg = f"Error loading Excel file: {str(e)}"
            print(error_msg)
            allure.attach(
                body=error_msg,
                name="Excel Error",
                attachment_type=allure.attachment_type.TEXT
            )
            raise

    @staticmethod
    def get_sheet(workbook, sheet_name):
        """Get sheet from workbook"""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        return workbook[sheet_name]

    @staticmethod
    def read_data(file_path, sheet_name, row_num, col_num):
        """Read data from specific cell"""
        try:
            workbook = ExcelHandler.load_workbook(file_path)
            sheet = ExcelHandler.get_sheet(workbook, sheet_name)
            return sheet.cell(row=row_num, column=col_num).value
        except Exception as e:
            print(f"Error reading cell data: {str(e)}")
            return None

    @staticmethod
    def get_row_count(file_path, sheet_name):
        """Get total number of rows in sheet"""
        try:
            workbook = ExcelHandler.load_workbook(file_path)
            sheet = ExcelHandler.get_sheet(workbook, sheet_name)
            return sheet.max_row
        except Exception as e:
            print(f"Error getting row count: {str(e)}")
            raise

    @staticmethod
    def generate_report(results):
        """Generate Excel report from test results"""
        try:
            # Prepare data for Excel
            data = []
            for index, result in enumerate(results, 1):
                data.append({
                    'S.no': index,
                    'URL': result['url'],
                    'Pass/Fail': 'Pass' if result['status'] == 'Success' else 'Fail',
                    'Time(ms)': round(result.get('load_time', 0), 2)
                })

            # Create DataFrame
            df = pd.DataFrame(data)

            # Generate timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Get output directory path
            output_dir = Configuration.get_path("output_excel")

            # Ensure output directory exists
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # Save Excel file
            output_path = os.path.join(output_dir, f"TestResults_{timestamp}.xlsx")

            # Create workbook and select active sheet
            wb = openpyxl.Workbook()
            ws = wb.active

            # Write headers
            headers = ['S.no', 'URL', 'Pass/Fail', 'Time(ms)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, (_, value) in enumerate(row_data.items(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply basic styling
            for cell in ws['1:1']:
                cell.style = 'Headline 1'

            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Save the workbook
            wb.save(output_path)

            print(f"Excel report generated: {output_path}")
            return output_path

        except Exception as e:
            error_msg = f"Error generating Excel report: {str(e)}"
            print(error_msg)
            allure.attach(
                body=error_msg,
                name="Report Generation Error",
                attachment_type=allure.attachment_type.TEXT
            )
            raise

    @staticmethod
    def backup_previous_report():
        """Backup previous Excel report if exists"""
        try:
            output_dir = Configuration.get_path("output_excel")
            backup_dir = Configuration.get_path("backup_excel")

            if not os.path.exists(output_dir) or not os.listdir(output_dir):
                return

            # Create backup directory if it doesn't exist
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Move each file in output directory to backup
            for filename in os.listdir(output_dir):
                if filename.endswith('.xlsx'):
                    src_path = os.path.join(output_dir, filename)
                    backup_filename = f"Backup_{timestamp}_{filename}"
                    dst_path = os.path.join(backup_dir, backup_filename)

                    try:
                        os.rename(src_path, dst_path)
                        print(f"Backed up report to: {dst_path}")
                    except Exception as e:
                        print(f"Error backing up {filename}: {str(e)}")

        except Exception as e:
            print(f"Error during backup: {str(e)}")
