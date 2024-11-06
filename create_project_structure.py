# create_project.py
import os
import json
import shutil
import pandas as pd


class ProjectCreator:
    def __init__(self):
        self.project_root = os.getcwd()

    def create_directory_structure(self):
        """Create or recreate all necessary directories"""
        directories = [
            'tests',
            'utils',
            'resources',
            'reports/screenshots',
            'reports/allure-results',
            'reports/ExtentReport',
            'reports/backup',
            'logs'
        ]

        print("\nCreating directory structure...")
        for directory in directories:
            dir_path = os.path.join(self.project_root, directory)
            if os.path.exists(dir_path):
                shutil.rmtree(dir_path)
            os.makedirs(dir_path)
            print(f"Created directory: {dir_path}")

    def create_sample_excel(self):
        """Create sample Excel file with test URLs"""
        print("\nCreating sample Excel file...")
        sample_urls = [
            'https://www.google.com',
            'https://www.github.com',
            'https://www.python.org',
            'https://www.microsoft.com',
            'https://www.apple.com',
            'https://thisisaninvalidurl.com'
        ]

        df = pd.DataFrame({'URL': sample_urls})
        excel_path = os.path.join(self.project_root, 'resources', 'links.xlsx')
        df.to_excel(excel_path, index=False, sheet_name='Sheet1')
        print(f"Created Excel file: {excel_path}")

    def create_config_json(self):
        """Create configuration file"""
        print("\nCreating config.json...")
        config = {
            "excel_path": "resources/links.xlsx",
            "chrome_driver_path": "",  # User needs to update this
            "sheet_name": "Sheet1",
            "page_load_timeout": 60,
            "max_retries": 2,
            "retry_delay": 2,
            "wait_between_urls": 2
        }

        with open('config.json', 'w') as f:
            json.dump(config, f, indent=4)
        print("Created config.json")

    def create_requirements(self):
        """Create requirements.txt"""
        print("\nCreating requirements.txt...")
        requirements = """
selenium>=4.16.0
webdriver-manager>=4.0.1
pytest>=7.4.3
allure-pytest>=2.13.2
openpyxl>=3.1.2
pytest-html>=4.1.1
requests>=2.31.0
urllib3>=2.0.7
pandas>=2.1.1
""".strip()

        with open('requirements.txt', 'w') as f:
            f.write(requirements)
        print("Created requirements.txt")

    def create_pytest_ini(self):
        """Create pytest.ini"""
        print("\nCreating pytest.ini...")
        content = """
[pytest]
pythonpath = .
testpaths = tests
python_files = test_*.py
python_classes = Test*
python_functions = test_*
addopts = -v -s --alluredir=reports/allure-results
""".strip()

        with open('pytest.ini', 'w') as f:
            f.write(content)
        print("Created pytest.ini")

    def create_utils_files(self):
        """Create utility files"""
        print("\nCreating utility files...")

        # Create __init__.py
        utils_init = os.path.join(self.project_root, 'utils', '__init__.py')
        with open(utils_init, 'w') as f:
            pass

        # Create config_handler.py
        config_handler = """
import os
import json
import shutil
import allure
from datetime import datetime

class Configuration:
    @staticmethod
    def load_config():
        config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'config.json')
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading config file: {str(e)}")
            return {}

    @staticmethod
    def get_config():
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        custom_config = Configuration.load_config()

        if 'excel_path' in custom_config and not os.path.isabs(custom_config['excel_path']):
            custom_config['excel_path'] = os.path.join(project_root, custom_config['excel_path'])

        default_config = {
            "excel_path": os.path.join(project_root, "resources", "links.xlsx"),
            "screenshots_dir": os.path.join(project_root, "reports", "screenshots"),
            "reports_dir": os.path.join(project_root, "reports", "allure-results"),
            "extent_report_dir": os.path.join(project_root, "reports", "ExtentReport"),
            "backup_dir": os.path.join(project_root, "reports", "backup"),
            "logs_dir": os.path.join(project_root, "logs"),
            "chrome_driver_path": custom_config.get("chrome_driver_path", ""),
            "sheet_name": "Sheet1",
            "page_load_timeout": 60,
            "max_retries": 2,
            "retry_delay": 2,
            "wait_between_urls": 2
        }

        default_config.update(custom_config)
        return default_config

    @staticmethod
    def get_path(path_name):
        config = Configuration.get_config()
        paths = {
            "excel": config["excel_path"],
            "screenshots": config["screenshots_dir"],
            "reports": config["reports_dir"],
            "extent_report": config["extent_report_dir"],
            "backup": config["backup_dir"],
            "logs": config["logs_dir"],
            "chrome_driver": config["chrome_driver_path"]
        }
        path = paths.get(path_name)
        return os.path.abspath(path) if path else None
"""
        with open(os.path.join(self.project_root, 'utils', 'config_handler.py'), 'w') as f:
            f.write(config_handler.strip())

        # Create excel_handler.py
        excel_handler = """
import openpyxl
import allure

class ExcelHandler:
    @staticmethod
    @allure.step("Loading Excel workbook")
    def load_workbook(file_path):
        try:
            return openpyxl.load_workbook(file_path)
        except Exception as e:
            allure.attach(
                body=f"Error loading Excel file: {str(e)}",
                name="Excel Error",
                attachment_type=allure.attachment_type.TEXT
            )
            raise

    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = ExcelHandler.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    @allure.step("Reading Excel data")
    def read_data(file, sheet_name, rownum, columnno):
        workbook = ExcelHandler.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=rownum, column=columnno).value
"""
        with open(os.path.join(self.project_root, 'utils', 'excel_handler.py'), 'w') as f:
            f.write(excel_handler.strip())

        # URLs will be processed using the previously provided web_handler.py code
        print("Created utility files")

    def create_test_files(self):
        """Create test files"""
        print("\nCreating test files...")

        # Create __init__.py
        test_init = os.path.join(self.project_root, 'tests', '__init__.py')
        with open(test_init, 'w') as f:
            pass

        # Create conftest.py
        conftest = """
import pytest
import allure
from datetime import datetime

@pytest.fixture(scope="session", autouse=True)
def setup_teardown():
    allure.attach(
        body="Test session started",
        name="Session Start",
        attachment_type=allure.attachment_type.TEXT
    )
    yield
    allure.attach(
        body="Test session ended",
        name="Session End",
        attachment_type=allure.attachment_type.TEXT
    )

@pytest.fixture(scope="function", autouse=True)
def test_case_setup(request):
    test_name = request.node.name
    allure.attach(
        body=f"Starting test: {test_name}",
        name="Test Start",
        attachment_type=allure.attachment_type.TEXT
    )
    start_time = datetime.now()
    yield
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    allure.attach(
        body=f"Test duration: {duration:.2f} seconds",
        name="Test Duration",
        attachment_type=allure.attachment_type.TEXT
    )
"""
        with open(os.path.join(self.project_root, 'tests', 'conftest.py'), 'w') as f:
            f.write(conftest.strip())

        print("Created test files")

    def create_all(self):
        """Create all project files and directories"""
        print("Starting project creation...")
        self.create_directory_structure()
        self.create_sample_excel()
        self.create_config_json()
        self.create_requirements()
        self.create_pytest_ini()
        self.create_utils_files()
        self.create_test_files()
        print("\nProject creation completed successfully!")
        print("\nNext steps:")
        print("1. Update chrome_driver_path in config.json")
        print("2. Install requirements: pip install -r requirements.txt")
        print("3. Run tests: pytest tests/test_url_processor.py -v")


if __name__ == "__main__":
    creator = ProjectCreator()
    creator.create_all()